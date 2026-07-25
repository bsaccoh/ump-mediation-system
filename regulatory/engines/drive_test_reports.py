import io
import csv
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


def generate_drive_test_pdf(campaign, analysis) -> bytes:
    """Generate official printable PDF Drive Test Audit Certificate."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18, leading=22, textColor=colors.HexColor('#1F4E78'))
    h2_style = ParagraphStyle('Heading2Style', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=12, leading=15, textColor=colors.HexColor('#1F4E78'))
    body_style = ParagraphStyle('BodyStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=12, textColor=colors.HexColor('#333333'))
    hdr_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, leading=11, textColor=colors.white)

    elements = []

    # Title & NatCA Regulatory Header
    elements.append(Paragraph("NATIONAL TELECOMMUNICATIONS AUTHORITY (NatCA)", title_style))
    elements.append(Paragraph("<b>Official Mobile Radio Network Drive Test Audit Certificate</b>", h2_style))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1F4E78'), spaceAfter=10))

    comp_status = "COMPLIANT" if getattr(analysis, 'natca_compliant', False) else "NON-COMPLIANT"
    comp_color = colors.HexColor('#28a745') if getattr(analysis, 'natca_compliant', False) else colors.HexColor('#dc3545')

    elements.append(Paragraph(f"<b>Campaign Name:</b> {campaign.name} &nbsp;|&nbsp; <b>Operator:</b> {campaign.operator_name.upper()}", body_style))
    elements.append(Paragraph(f"<b>Survey Date:</b> {campaign.test_date} &nbsp;|&nbsp; <b>Technology:</b> {campaign.technology} &nbsp;|&nbsp; <b>Tool:</b> {campaign.tool_used}", body_style))
    elements.append(Paragraph(f"<b>Region / Route:</b> {campaign.region} ({campaign.route_description or 'Standard Route'})", body_style))
    elements.append(Spacer(1, 10))

    # Executive Audit Summary Table
    elements.append(Paragraph("1. Executive Field Quality Summary & NatCA Compliance", h2_style))

    summary_data = [
        [Paragraph("Metric", hdr_style), Paragraph("Measured Value", hdr_style), Paragraph("NatCA Target Rule", hdr_style), Paragraph("Compliance Status", hdr_style)],
        [Paragraph("Coverage % (RSRP >= -110 dBm)", body_style), Paragraph(f"{analysis.coverage_pct}%", body_style), Paragraph(">= 90.00%", body_style), Paragraph("<b>PASS</b>" if analysis.coverage_pct >= Decimal('90') else "<b>FAIL</b>", body_style)],
        [Paragraph("Average RSRP (Signal Strength)", body_style), Paragraph(f"{analysis.avg_rsrp} dBm", body_style), Paragraph(">= -105.00 dBm", body_style), Paragraph("<b>PASS</b>" if analysis.avg_rsrp >= Decimal('-105') else "<b>FAIL</b>", body_style)],
        [Paragraph("Average DL Throughput", body_style), Paragraph(f"{analysis.avg_dl_throughput} Mbps", body_style), Paragraph(">= 5.00 Mbps", body_style), Paragraph("<b>PASS</b>" if analysis.avg_dl_throughput >= Decimal('5') else "<b>FAIL</b>", body_style)],
        [Paragraph("Call Setup Success Rate (CSSR)", body_style), Paragraph(f"{analysis.cssr_pct}%", body_style), Paragraph(">= 95.00%", body_style), Paragraph("<b>PASS</b>" if analysis.cssr_pct >= Decimal('95') else "<b>FAIL</b>", body_style)],
        [Paragraph("Call Drop Rate (CDR)", body_style), Paragraph(f"{analysis.drop_rate_pct}%", body_style), Paragraph("<= 2.00%", body_style), Paragraph("<b>PASS</b>" if analysis.drop_rate_pct <= Decimal('2') else "<b>FAIL</b>", body_style)],
    ]

    t1 = Table(summary_data, colWidths=[180, 110, 120, 110])
    t1.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(t1)
    elements.append(Spacer(1, 15))

    # RF Percentiles Table
    elements.append(Paragraph("2. Radio Frequency (RF) Percentiles Breakdown", h2_style))
    rsrp_pcts = getattr(analysis, 'rsrp_percentiles', {}) or {}
    sinr_pcts = getattr(analysis, 'sinr_percentiles', {}) or {}
    tp_pcts = getattr(analysis, 'tp_percentiles', {}) or {}

    pct_data = [
        [Paragraph("RF Metric", hdr_style), Paragraph("p5", hdr_style), Paragraph("p25", hdr_style), Paragraph("p50 (Median)", hdr_style), Paragraph("p90", hdr_style), Paragraph("p95", hdr_style)],
        [Paragraph("RSRP (dBm)", body_style), Paragraph(str(rsrp_pcts.get('p5', '-')), body_style), Paragraph(str(rsrp_pcts.get('p25', '-')), body_style), Paragraph(str(rsrp_pcts.get('p50', '-')), body_style), Paragraph(str(rsrp_pcts.get('p90', '-')), body_style), Paragraph(str(rsrp_pcts.get('p95', '-')), body_style)],
        [Paragraph("SINR (dB)", body_style), Paragraph(str(sinr_pcts.get('p5', '-')), body_style), Paragraph(str(sinr_pcts.get('p25', '-')), body_style), Paragraph(str(sinr_pcts.get('p50', '-')), body_style), Paragraph(str(sinr_pcts.get('p90', '-')), body_style), Paragraph(str(sinr_pcts.get('p95', '-')), body_style)],
        [Paragraph("DL Throughput (Mbps)", body_style), Paragraph(str(tp_pcts.get('p5', '-')), body_style), Paragraph(str(tp_pcts.get('p25', '-')), body_style), Paragraph(str(tp_pcts.get('p50', '-')), body_style), Paragraph(str(tp_pcts.get('p90', '-')), body_style), Paragraph(str(tp_pcts.get('p95', '-')), body_style)],
    ]

    t2 = Table(pct_data, colWidths=[140, 75, 75, 85, 75, 70])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(t2)

    doc.build(elements)
    return output.getvalue()


def generate_drive_test_excel(campaign, analysis) -> bytes:
    """Generate multi-tab Excel workbook for drive test campaign."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Executive Summary"

    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    ws1.append(["NATIONAL TELECOMMUNICATIONS AUTHORITY (NatCA)"])
    ws1.append(["Official Mobile Radio Network Drive Test Quality Audit Workbook"])
    ws1.append([f"Campaign: {campaign.name} | Operator: {campaign.operator_name} | Date: {campaign.test_date}"])
    ws1.append([])

    ws1['A1'].font = title_font

    ws1.append(["Metric Name", "Measured Value", "Target Rule", "Audit Status"])
    for col_idx in range(1, 5):
        cell = ws1.cell(row=5, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    ws1.append(["Total Samples", analysis.total_samples, "-", "-"])
    ws1.append(["Coverage % (RSRP >= -110 dBm)", f"{analysis.coverage_pct}%", ">= 90.00%", "PASS" if analysis.coverage_pct >= Decimal('90') else "FAIL"])
    ws1.append(["Average RSRP", f"{analysis.avg_rsrp} dBm", ">= -105.00 dBm", "PASS" if analysis.avg_rsrp >= Decimal('-105') else "FAIL"])
    ws1.append(["Average SINR", f"{analysis.avg_sinr} dB", ">= 3.00 dB", "PASS" if analysis.avg_sinr >= Decimal('3') else "FAIL"])
    ws1.append(["Average DL Throughput", f"{analysis.avg_dl_throughput} Mbps", ">= 5.00 Mbps", "PASS" if analysis.avg_dl_throughput >= Decimal('5') else "FAIL"])

    # Sheet 2: Raw Samples
    ws2 = wb.create_sheet(title="Raw Field Samples")
    ws2.append(["Sample ID", "Latitude", "Longitude", "RSRP (dBm)", "SINR (dB)", "DL Throughput (Mbps)", "CSSR Status", "Drop Status", "Voice MOS", "Blackspot Flag"])
    for col_idx in range(1, 11):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    for s in campaign.samples.all()[:5000]:
        ws2.append([
            s.pk,
            float(s.latitude),
            float(s.longitude),
            float(s.rsrp) if s.rsrp is not None else None,
            float(s.sinr) if s.sinr is not None else None,
            float(s.dl_throughput) if s.dl_throughput is not None else None,
            s.cssr_status,
            s.drop_status,
            float(s.voice_mos) if s.voice_mos is not None else None,
            s.is_blackspot,
        ])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_drive_test_csv(campaign) -> str:
    """Generate GIS-ready CSV export for QGIS / ArcGIS."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'sample_id', 'latitude', 'longitude', 'timestamp',
        'rsrp_dbm', 'rsrq_db', 'sinr_db', 'dl_tp_mbps', 'ul_tp_mbps',
        'cssr_status', 'drop_status', 'voice_mos', 'technology', 'pci', 'is_blackspot'
    ])

    for s in campaign.samples.all():
        writer.writerow([
            s.pk,
            float(s.latitude),
            float(s.longitude),
            s.timestamp.isoformat() if s.timestamp else '',
            float(s.rsrp) if s.rsrp is not None else '',
            float(s.rsrq) if s.rsrq is not None else '',
            float(s.sinr) if s.sinr is not None else '',
            float(s.dl_throughput) if s.dl_throughput is not None else '',
            float(s.ul_throughput) if s.ul_throughput is not None else '',
            s.cssr_status if s.cssr_status is not None else '',
            s.drop_status if s.drop_status is not None else '',
            float(s.voice_mos) if s.voice_mos is not None else '',
            s.technology,
            s.pci if s.pci is not None else '',
            1 if s.is_blackspot else 0,
        ])

    return output.getvalue()
