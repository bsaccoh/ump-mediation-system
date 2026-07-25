import io
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


def generate_natca_audit_excel(audit_data: dict) -> bytes:
    """Generate multi-tab Excel Workbook for NatCA Regulatory Compliance Audit."""
    wb = openpyxl.Workbook()

    # Define Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="595959")
    bold_font = Font(name="Calibri", size=11, bold=True)

    border_thin = Side(style='thin', color='D9D9D9')
    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    meta = audit_data.get('audit_metadata', {})
    summaries = audit_data.get('operator_summaries', [])
    rankings = audit_data.get('district_rankings', [])
    violations = audit_data.get('kpi_violations', [])

    # ---------------------------------------------------------
    # Sheet 1: Operator Compliance & Penalty Summary
    # ---------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Executive Audit Summary"

    ws1.append(["NATIONAL TELECOMMUNICATIONS AUTHORITY (NatCA)"])
    ws1.append(["Official Telecom Regulatory Quality of Service Audit & SLA Penalty Report"])
    ws1.append([f"Audit Stamp: {meta.get('audit_stamp')} | Generated: {meta.get('generated_at')}"])
    ws1.append([])

    ws1['A1'].font = title_font
    ws1['A2'].font = Font(name="Calibri", size=12, bold=True, color="333333")
    ws1['A3'].font = subtitle_font

    headers1 = ["Operator Code", "Total Audited Measurements", "Compliant Count", "SLA Breach Count", "Compliance Score (%)", "Assessed Penalty Fee ($)", "Audit Status"]
    ws1.append(headers1)
    for col_idx in range(1, len(headers1) + 1):
        cell = ws1.cell(row=5, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for op in summaries:
        row_vals = [
            op['operator_code'].upper(),
            op['total_measurements'],
            op['compliant_measurements'],
            op['breach_count'],
            f"{op['compliance_score']}%",
            f"${float(op['total_penalty_fee']):,.2f}",
            op['status']
        ]
        ws1.append(row_vals)

    ws1.append([])
    ws1.append(["System Totals:", meta.get('total_system_entries'), "-", meta.get('total_system_breaches'), "-", f"${float(meta.get('total_penalties_assessed', 0)):,.2f}", "-"])
    last_row = ws1.max_row
    for c in range(1, 8):
        ws1.cell(row=last_row, column=c).font = bold_font

    # ---------------------------------------------------------
    # Sheet 2: District Quality Rankings
    # ---------------------------------------------------------
    ws2 = wb.create_sheet(title="District Rankings")
    ws2.append(["Sierra Leone District Network Quality Audit Rankings"])
    ws2['A1'].font = title_font
    ws2.append([])

    headers2 = ["District Name", "Province / Region", "Total Audited", "Compliant Measurements", "Compliance Rate (%)", "Quality Grade"]
    ws2.append(headers2)
    for col_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    for d in rankings:
        ws2.append([
            d['district'],
            d['region'],
            d['total_audited'],
            d['pass_count'],
            f"{d['compliance_rate']}%",
            d['quality_grade']
        ])

    # ---------------------------------------------------------
    # Sheet 3: Detailed SLA Violations Log
    # ---------------------------------------------------------
    ws3 = wb.create_sheet(title="SLA Violations Log")
    ws3.append(["Detailed KPI SLA Violations & Penalty Breakdown"])
    ws3['A1'].font = title_font
    ws3.append([])

    headers3 = ["Event ID", "Operator", "KPI Code", "KPI Name", "Period Date", "Region / District", "Cell ID", "Measured Value", "NatCA Target", "Penalty ($)", "Severity"]
    ws3.append(headers3)
    for col_idx in range(1, len(headers3) + 1):
        cell = ws3.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    for v in violations:
        ws3.append([
            v['id'],
            v['operator_code'].upper(),
            v['kpi_code'],
            v['kpi_name'],
            v['period_date'],
            f"{v['region']} / {v['district']}",
            v['cell_id'],
            v['measured_value'],
            v['target_threshold'],
            f"${float(v['penalty_fee']):,.2f}",
            v['severity']
        ])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_natca_audit_pdf(audit_data: dict) -> bytes:
    """Generate printable official PDF Executive Summary Audit Report."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18, leading=22, textColor=colors.HexColor('#1F4E78'))
    h2_style = ParagraphStyle('Heading2Style', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=13, leading=16, textColor=colors.HexColor('#1F4E78'))
    body_style = ParagraphStyle('BodyStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=12, textColor=colors.HexColor('#333333'))
    hdr_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, leading=11, textColor=colors.white)

    meta = audit_data.get('audit_metadata', {})
    summaries = audit_data.get('operator_summaries', [])
    rankings = audit_data.get('district_rankings', [])

    elements = []

    # Title & NatCA Regulatory Header
    elements.append(Paragraph("NATIONAL TELECOMMUNICATIONS AUTHORITY (NatCA)", title_style))
    elements.append(Paragraph("<b>Official Telecom Quality of Service (QoS) Compliance Audit Report</b>", h2_style))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1F4E78'), spaceAfter=10))

    elements.append(Paragraph(f"<b>Audit Reference:</b> {meta.get('audit_stamp')} &nbsp;|&nbsp; <b>Generated:</b> {meta.get('generated_at')}", body_style))
    elements.append(Paragraph(f"<b>Total Audited Measurements:</b> {meta.get('total_system_entries', 0):,} &nbsp;|&nbsp; <b>Assessed Penalty Fees:</b> ${float(meta.get('total_penalties_assessed', 0)):,.2f}", body_style))
    elements.append(Spacer(1, 15))

    # Table 1: Operator Compliance & Penalty Summary
    elements.append(Paragraph("1. Multi-Operator SLA Compliance & Penalty Assessment", h2_style))

    table_data = [[
        Paragraph("Operator", hdr_style),
        Paragraph("Audited Entries", hdr_style),
        Paragraph("Breach Count", hdr_style),
        Paragraph("Compliance Score", hdr_style),
        Paragraph("Penalty Fee ($)", hdr_style),
        Paragraph("Audit Status", hdr_style),
    ]]

    for op in summaries:
        table_data.append([
            Paragraph(op['operator_code'].upper(), body_style),
            Paragraph(str(op['total_measurements']), body_style),
            Paragraph(str(op['breach_count']), body_style),
            Paragraph(f"{op['compliance_score']}%", body_style),
            Paragraph(f"${float(op['total_penalty_fee']):,.2f}", body_style),
            Paragraph(f"<b>{op['status']}</b>", body_style),
        ])

    t1 = Table(table_data, colWidths=[80, 85, 80, 100, 100, 85])
    t1.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(t1)
    elements.append(Spacer(1, 15))

    # Table 2: District Quality Rankings
    elements.append(Paragraph("2. Sierra Leone District Network Quality Rankings", h2_style))
    t2_data = [[
        Paragraph("District", hdr_style),
        Paragraph("Province / Region", hdr_style),
        Paragraph("Audited Count", hdr_style),
        Paragraph("Compliance Rate", hdr_style),
        Paragraph("Quality Grade", hdr_style),
    ]]

    for d in rankings[:10]:
        t2_data.append([
            Paragraph(d['district'], body_style),
            Paragraph(d['region'], body_style),
            Paragraph(str(d['total_audited']), body_style),
            Paragraph(f"{d['compliance_rate']}%", body_style),
            Paragraph(f"<b>{d['quality_grade']}</b>", body_style),
        ])

    t2 = Table(t2_data, colWidths=[120, 130, 90, 100, 90])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(t2)

    doc.build(elements)
    return output.getvalue()
