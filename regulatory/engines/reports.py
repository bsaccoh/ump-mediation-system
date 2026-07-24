"""NATCOM periodic-report engine.

Four report types::

  * TRAFFIC               — voice/SMS/data totals across MSC, IMS, PGW/SGSN/SGW
  * REVENUE               — interconnect INBOUND vs OUTBOUND + retail (if entered)
  * SUBSCRIBER            — distinct active IMSIs (CDR-derived proxy)
  * INTERCONNECT_SUMMARY  — partner-by-partner traffic + settlement status

Each ``generate_*_report(start, end)`` returns a JSON-serialisable dict that
both the PDF and Excel renderers consume.
"""
from __future__ import annotations

import io
from datetime import date, datetime, time, timedelta
from decimal import Decimal

from django.core.files.base import ContentFile
from django.db.models import Sum, Count
from django.utils import timezone

from core.models import AuditLog

from ..models import (
    RegulatoryProfile, RegulatoryReport, RetailRevenue,
)


# ---------------------------------------------------------------------------
# 1. Traffic report
# ---------------------------------------------------------------------------

def generate_traffic_report(start: date, end: date) -> dict:
    from streams.msc.models import MSCRecord
    from streams.ims.models import IMSRecord
    from streams.pgw.models import PGWRecord
    from streams.sgsn.models import SGSNRecord
    from streams.sgw.models import SGWRecord

    s = datetime.combine(start, time.min)
    e = datetime.combine(end + timedelta(days=1), time.min)

    msc_agg = MSCRecord.objects.filter(start_time__gte=s, start_time__lt=e).values('record_type').annotate(
        count=Count('id'), duration=Sum('duration'),
    )
    ims_agg = IMSRecord.objects.filter(start_time__gte=s, start_time__lt=e).aggregate(
        calls=Count('id'), duration=Sum('duration'),
    )
    # PGW/SGSN/SGW data_volume_* are CharFields; sum manually with a safe int parse
    pgw_sessions = PGWRecord.objects.filter(start_time__gte=s, start_time__lt=e).count()
    sgsn_sessions = SGSNRecord.objects.filter(start_time__gte=s, start_time__lt=e).count()
    sgw_sessions = SGWRecord.objects.filter(start_time__gte=s, start_time__lt=e).count()
    total_bytes = _sum_data_volumes_safe(
        PGWRecord.objects.filter(start_time__gte=s, start_time__lt=e),
        SGSNRecord.objects.filter(start_time__gte=s, start_time__lt=e),
        SGWRecord.objects.filter(start_time__gte=s, start_time__lt=e),
    )

    voice_calls = sum(r['count'] for r in msc_agg if r['record_type'] in ('MOC', 'MTC'))
    voice_minutes = sum((r['duration'] or 0) for r in msc_agg if r['record_type'] in ('MOC', 'MTC')) / 60.0
    sms_count = sum(r['count'] for r in msc_agg if r['record_type'] in ('SMSMO', 'SMSMT'))
    gw_count = sum(r['count'] for r in msc_agg if r['record_type'] in ('GWO', 'GWI', 'GWOUT', 'GWIN'))

    return {
        'period_start': start.isoformat(),
        'period_end': end.isoformat(),
        'voice_calls_msc': voice_calls,
        'voice_minutes_msc': round(voice_minutes, 2),
        'voice_calls_ims': ims_agg.get('calls') or 0,
        'voice_minutes_ims': round((ims_agg.get('duration') or 0) / 60.0, 2),
        'sms_count': sms_count,
        'gateway_legs': gw_count,
        'data_sessions_pgw': pgw_sessions,
        'data_sessions_sgsn': sgsn_sessions,
        'data_sessions_sgw': sgw_sessions,
        'data_gb_total': round(total_bytes / (1024**3), 3),
    }


def _sum_data_volumes_safe(*querysets) -> int:
    """Sum data_volume_up + data_volume_down across given querysets.

    PGW/SGSN/SGW now store these as BigIntegerField — single SQL SUM each.
    """
    total = 0
    for qs in querysets:
        agg = qs.aggregate(up=Sum('data_volume_up'), dn=Sum('data_volume_down'))
        total += int(agg.get('up') or 0) + int(agg.get('dn') or 0)
    return total


# ---------------------------------------------------------------------------
# 2. Revenue report
# ---------------------------------------------------------------------------

def generate_revenue_report(start: date, end: date) -> dict:
    from interconnect.models import Invoice

    base = (Invoice.objects.exclude(status=Invoice.Status.VOID)
            .filter(billing_cycle__period_end__gte=start,
                     billing_cycle__period_end__lte=end))
    inbound = base.filter(direction='INBOUND').aggregate(
        s=Sum('total_local'))['s'] or Decimal('0')
    outbound = base.filter(direction='OUTBOUND').aggregate(
        s=Sum('total_local'))['s'] or Decimal('0')
    inbound_count = base.filter(direction='INBOUND').count()
    outbound_count = base.filter(direction='OUTBOUND').count()

    # Retail (sum across all months that intersect the window)
    retail_rows = []
    months = _month_range(start, end)
    for y, m in months:
        try:
            r = RetailRevenue.objects.get(period_year=y, period_month=m)
            retail_rows.append({
                'period': f'{y}-{m:02d}',
                'voice': float(r.voice_revenue),
                'sms': float(r.sms_revenue),
                'data': float(r.data_revenue),
                'other': float(r.other_revenue),
                'total': float(r.total),
                'currency': r.currency,
            })
        except RetailRevenue.DoesNotExist:
            continue
    retail_total = sum(Decimal(str(r['total'])) for r in retail_rows)

    profile = RegulatoryProfile.get_or_create_default()

    return {
        'period_start': start.isoformat(),
        'period_end': end.isoformat(),
        'interconnect_inbound': float(inbound),
        'interconnect_outbound': float(outbound),
        'interconnect_net': float(inbound - outbound),
        'inbound_count': inbound_count,
        'outbound_count': outbound_count,
        'retail_total': float(retail_total),
        'retail_rows': retail_rows,
        'gross_revenue': float(inbound + retail_total),
        'currency': profile.home_currency,
    }


def _month_range(start: date, end: date):
    months = []
    cursor = date(start.year, start.month, 1)
    while cursor <= end:
        months.append((cursor.year, cursor.month))
        if cursor.month == 12:
            cursor = date(cursor.year + 1, 1, 1)
        else:
            cursor = date(cursor.year, cursor.month + 1, 1)
    return months


# ---------------------------------------------------------------------------
# 3. Subscriber report — distinct IMSIs as a proxy for active subscribers
# ---------------------------------------------------------------------------

def generate_subscriber_report(start: date, end: date) -> dict:
    from streams.msc.models import MSCRecord
    from streams.ims.models import IMSRecord
    from streams.pgw.models import PGWRecord

    s = datetime.combine(start, time.min)
    e = datetime.combine(end + timedelta(days=1), time.min)

    msc_imsis = MSCRecord.objects.filter(
        start_time__gte=s, start_time__lt=e,
    ).exclude(imsi='').values('imsi').distinct().count()
    ims_imsis = IMSRecord.objects.filter(
        start_time__gte=s, start_time__lt=e,
    ).exclude(imsi='').values('imsi').distinct().count()
    pgw_imsis = PGWRecord.objects.filter(
        start_time__gte=s, start_time__lt=e,
    ).exclude(imsi='').values('imsi').distinct().count()

    # Prepaid / postpaid split — prepaid_flag is a CharField with canonical
    # 'PREPAID' / 'POSTPAID' values (set by the CAMEL-IN-trigger rule in the
    # MSC processor; see core.utils.prepaid.derive_msc_prepaid_flag).
    prepaid = MSCRecord.objects.filter(
        start_time__gte=s, start_time__lt=e, prepaid_flag='PREPAID',
    ).exclude(imsi='').values('imsi').distinct().count()
    postpaid = MSCRecord.objects.filter(
        start_time__gte=s, start_time__lt=e, prepaid_flag='POSTPAID',
    ).exclude(imsi='').values('imsi').distinct().count()

    return {
        'period_start': start.isoformat(),
        'period_end': end.isoformat(),
        'distinct_imsi_msc': msc_imsis,
        'distinct_imsi_ims': ims_imsis,
        'distinct_imsi_data': pgw_imsis,
        'prepaid_subscribers': prepaid,
        'postpaid_subscribers': postpaid,
        'note': 'Subscriber counts are CDR-derived (active subscribers in period). For SIM-registry counts, integrate with the subscriber DB.',
    }


# ---------------------------------------------------------------------------
# 4. Interconnect summary
# ---------------------------------------------------------------------------

def generate_interconnect_summary(start: date, end: date) -> dict:
    from interconnect.models import Invoice, BillingCycle

    cycles = (BillingCycle.objects.select_related('partner')
              .filter(period_end__gte=start, period_start__lte=end))
    rows = []
    for c in cycles:
        invs = c.invoices.exclude(status=Invoice.Status.VOID)
        inbound = invs.filter(direction='INBOUND').aggregate(s=Sum('total_local'))['s'] or Decimal('0')
        outbound = invs.filter(direction='OUTBOUND').aggregate(s=Sum('total_local'))['s'] or Decimal('0')
        paid = sum(float(i.amount_paid) for i in invs)
        outstanding = sum(float(i.amount_outstanding) for i in invs)
        rows.append({
            'partner': c.partner.code, 'partner_name': c.partner.name,
            'period': f'{c.period_start}..{c.period_end}',
            'voice_minutes': float(c.our_voice_minutes),
            'voice_calls': c.our_voice_calls,
            'sms': c.our_sms,
            'data_mb': float(c.our_data_mb),
            'inbound': float(inbound), 'outbound': float(outbound),
            'paid': paid, 'outstanding': outstanding,
        })
    return {
        'period_start': start.isoformat(),
        'period_end': end.isoformat(),
        'rows': rows,
        'total_inbound': sum(r['inbound'] for r in rows),
        'total_outbound': sum(r['outbound'] for r in rows),
        'total_outstanding': sum(r['outstanding'] for r in rows),
    }


# ---------------------------------------------------------------------------
# Top-level entry: generate_report dispatches + persists
# ---------------------------------------------------------------------------

GENERATORS = {
    'TRAFFIC':               generate_traffic_report,
    'REVENUE':               generate_revenue_report,
    'SUBSCRIBER':            generate_subscriber_report,
    'INTERCONNECT_SUMMARY':  generate_interconnect_summary,
}


def generate_report(report_type: str, start: date, end: date, user=None) -> RegulatoryReport:
    if report_type not in GENERATORS:
        raise ValueError(f'Unknown report_type {report_type!r}')
    payload = GENERATORS[report_type](start, end)

    report = RegulatoryReport.objects.create(
        report_type=report_type,
        period_start=start, period_end=end,
        status=RegulatoryReport.Status.DRAFT,
        summary_json=payload,
        generated_by=user if user and user.is_authenticated else None,
    )

    # Render & attach artefacts (best-effort)
    try:
        pdf_bytes = render_pdf(payload, report_type)
        report.pdf_file.save(f'natcom_{report_type}_{start}_{end}.pdf',
                              ContentFile(pdf_bytes), save=False)
    except Exception:
        pass
    try:
        xlsx_bytes = render_xlsx(payload, report_type)
        report.xlsx_file.save(f'natcom_{report_type}_{start}_{end}.xlsx',
                                ContentFile(xlsx_bytes), save=False)
    except Exception:
        pass
    report.save(update_fields=['pdf_file', 'xlsx_file'])

    # Audit
    try:
        AuditLog.objects.create(
            user=user if user and user.is_authenticated else None,
            action='REGULATORY_REPORT_GENERATED',
            entity_type='RegulatoryReport',
            entity_id=str(report.pk),
            description=f'{report_type} {start}..{end}',
        )
    except Exception:
        pass

    return report


# ---------------------------------------------------------------------------
# PDF renderer (reportlab platypus)
# ---------------------------------------------------------------------------

def render_pdf(payload: dict, report_type: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=18*mm, rightMargin=18*mm,
                             topMargin=18*mm, bottomMargin=18*mm,
                             title=f'NATCOM {report_type}')
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle('h1', parent=styles['Heading1'], fontSize=16,
                          spaceAfter=2, textColor=colors.HexColor('#003366'))
    h2 = ParagraphStyle('h2', parent=styles['Heading3'], fontSize=11,
                          spaceAfter=2)
    small = ParagraphStyle('small', parent=styles['Normal'], fontSize=8.5)
    normal = styles['Normal']

    profile = RegulatoryProfile.get_or_create_default()

    story = []
    story.append(Paragraph('NATCOM REGULATORY RETURN', h1))
    story.append(Paragraph(
        f'<b>{_friendly_type(report_type)}</b> — period {payload["period_start"]} to {payload["period_end"]}',
        h2,
    ))
    story.append(Paragraph(f'Operator: Orange Sierra Leone &nbsp;&nbsp;|&nbsp;&nbsp; '
                           f'Regulator: {profile.regulator_name} &nbsp;&nbsp;|&nbsp;&nbsp; '
                           f'Generated: {timezone.now():%Y-%m-%d %H:%M}', small))
    story.append(Spacer(1, 6*mm))

    # Body table — derived from payload
    rows = _pdf_rows(payload, report_type)
    tbl = Table(rows, colWidths=[80*mm, 80*mm])
    tbl.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E6EEF7')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F7F7')]),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 10*mm))

    if 'note' in payload:
        story.append(Paragraph(f'<i>{payload["note"]}</i>', small))
        story.append(Spacer(1, 4*mm))

    story.append(Paragraph(
        f'This report is generated automatically from the operator\'s mediation system. '
        f'Levy: {profile.levy_pct}% · USF: {profile.usf_pct}% · Currency: {profile.home_currency}.',
        small,
    ))

    doc.build(story)
    return buf.getvalue()


def _friendly_type(t: str) -> str:
    return dict(RegulatoryReport.ReportType.choices).get(t, t)


def _pdf_rows(payload: dict, report_type: str):
    rows = [['Metric', 'Value']]
    SKIP = {'period_start', 'period_end', 'note', 'retail_rows', 'rows', 'currency'}
    for k, v in payload.items():
        if k in SKIP:
            continue
        if isinstance(v, float):
            v = f'{v:,.3f}'.rstrip('0').rstrip('.')
        rows.append([k.replace('_', ' ').title(), str(v)])
    # Inline tables for retail / per-partner detail
    if report_type == 'REVENUE' and payload.get('retail_rows'):
        rows.append(['—', '—'])
        rows.append(['Retail period', 'Total'])
        for r in payload['retail_rows']:
            rows.append([r['period'], f'{r["total"]:,.2f} {r["currency"]}'])
    if report_type == 'INTERCONNECT_SUMMARY' and payload.get('rows'):
        rows.append(['—', '—'])
        rows.append(['Partner', 'Inbound / Outbound / Outstanding'])
        for r in payload['rows'][:30]:
            rows.append([r['partner'],
                          f'{r["inbound"]:,.2f} / {r["outbound"]:,.2f} / {r["outstanding"]:,.2f}'])
    return rows


# ---------------------------------------------------------------------------
# Excel renderer (openpyxl)
# ---------------------------------------------------------------------------

def render_xlsx(payload: dict, report_type: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type[:31]

    bold = Font(bold=True)
    hdr_fill = PatternFill(start_color='E6EEF7', end_color='E6EEF7', fill_type='solid')

    ws.append([f'NATCOM Regulatory Return — {_friendly_type(report_type)}'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'Period: {payload["period_start"]} to {payload["period_end"]}'])
    ws.append([f'Generated: {timezone.now():%Y-%m-%d %H:%M}'])
    ws.append([])

    # Headline KPIs section
    ws.append(['Metric', 'Value'])
    for cell in ws[ws.max_row]:
        cell.font = bold
        cell.fill = hdr_fill

    SKIP = {'period_start', 'period_end', 'note', 'retail_rows', 'rows', 'currency'}
    for k, v in payload.items():
        if k in SKIP:
            continue
        ws.append([k.replace('_', ' ').title(), v])

    if report_type == 'REVENUE' and payload.get('retail_rows'):
        ws.append([])
        ws.append(['Retail revenue detail'])
        ws[ws.max_row][0].font = bold
        ws.append(['Period', 'Voice', 'SMS', 'Data', 'Other', 'Total', 'Currency'])
        for cell in ws[ws.max_row]:
            cell.font = bold
            cell.fill = hdr_fill
        for r in payload['retail_rows']:
            ws.append([r['period'], r['voice'], r['sms'], r['data'], r['other'],
                        r['total'], r['currency']])

    if report_type == 'INTERCONNECT_SUMMARY' and payload.get('rows'):
        ws.append([])
        ws.append(['Per-partner breakdown'])
        ws[ws.max_row][0].font = bold
        ws.append(['Partner', 'Name', 'Period', 'Voice min', 'Voice calls',
                    'SMS', 'Data MB', 'Inbound', 'Outbound', 'Paid', 'Outstanding'])
        for cell in ws[ws.max_row]:
            cell.font = bold
            cell.fill = hdr_fill
        for r in payload['rows']:
            ws.append([r['partner'], r['partner_name'], r['period'],
                        r['voice_minutes'], r['voice_calls'], r['sms'],
                        r['data_mb'], r['inbound'], r['outbound'],
                        r['paid'], r['outstanding']])

    # Auto-size first two columns
    for col in ('A', 'B'):
        max_len = max((len(str(c.value)) for c in ws[col] if c.value is not None), default=10)
        ws.column_dimensions[col].width = min(max_len + 2, 60)

    if 'note' in payload:
        ws.append([])
        ws.append([payload['note']])
        ws[ws.max_row][0].alignment = Alignment(wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
