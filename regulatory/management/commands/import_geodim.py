"""Import Cell Sites & Geo Dimension from official Geo-Dimension_Newsites_Upgrade_2026_V1.xlsx guide.

Usage:
    python manage.py import_geodim [--file path/to/file.xlsx]
"""
import os
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from django.core.management.base import BaseCommand
from django.db import transaction

from regulatory.models import NetworkCellSite, NetworkSectorCell


def _dec(val):
    if val in (None, ''):
        return None
    try:
        return Decimal(str(val).strip())
    except (InvalidOperation, ValueError, TypeError):
        return None


def _date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if not val:
        return None
    try:
        return datetime.strptime(str(val).strip()[:10], '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


class Command(BaseCommand):
    help = 'Import Cell Sites & Geo Dimensions from official Excel workbook (idempotent).'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            default='Geo-Dimension_Newsites_Upgrade_2026_V1.xlsx',
            help='Path to Geo-Dimension Excel spreadsheet',
        )

    def handle(self, *args, **opts):
        file_path = opts['file']
        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f'File not found: {file_path}'))
            return

        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sites_created = 0
        cells_created = 0

        # 1. Parse Physical Sites Sheet ('OSL_Physical Sites')
        if 'OSL_Physical Sites' in wb.sheetnames:
            sheet1 = wb['OSL_Physical Sites']
            rows1 = list(sheet1.iter_rows(values_only=True))
            if len(rows1) > 1:
                headers1 = [str(h).strip() if h else '' for h in rows1[0]]
                sites_to_upsert = []

                for r in rows1[1:]:
                    row_dict = dict(zip(headers1, r))
                    site_id = str(row_dict.get('SITE ID') or '').strip()
                    if not site_id:
                        continue

                    site_name = str(row_dict.get('SITE NAME') or site_id).strip()
                    lat = _dec(row_dict.get('LATITUDE'))
                    lng = _dec(row_dict.get('LONGITUDE'))
                    h_m = _dec(row_dict.get('Tower Height'))
                    tech = str(row_dict.get('Technology') or '2G3G4G').strip()
                    classif = str(row_dict.get('Classification') or '').strip()
                    natca_class = str(row_dict.get('NAtCa Sites Classification') or '').strip()
                    owner = str(row_dict.get('OWNER') or '').strip()
                    region = str(row_dict.get('Region') or 'Western Area').strip()
                    district = str(row_dict.get('District') or '').strip()
                    chiefdom = str(row_dict.get('Chiefdom') or '').strip()
                    location = str(row_dict.get('Location') or '').strip()
                    on_air = _date(row_dict.get('OnAir Date'))
                    site_type = str(row_dict.get('Site Type') or '').strip()

                    sites_to_upsert.append({
                        'operator_code': 'orange',
                        'site_id': site_id,
                        'cell_id': '',
                        'defaults': {
                            'site_name': site_name,
                            'technology': tech,
                            'classification': classif,
                            'natca_classification': natca_class,
                            'site_owner': owner,
                            'region': region,
                            'district': district,
                            'chiefdom': chiefdom,
                            'location': location,
                            'latitude': lat,
                            'longitude': lng,
                            'height_m': h_m,
                            'on_air_date': on_air,
                            'site_type': site_type,
                            'status': NetworkCellSite.Status.ACTIVE,
                        }
                    })

                with transaction.atomic():
                    for item in sites_to_upsert:
                        NetworkCellSite.objects.update_or_create(
                            operator_code=item['operator_code'],
                            site_id=item['site_id'],
                            cell_id=item['cell_id'],
                            defaults=item['defaults'],
                        )
                        sites_created += 1

                self.stdout.write(f'Imported/updated {sites_created} physical sites.')

        # 2. Parse Sector Cells Sheet ('GEO-DIM 2G_3G_4G_5G')
        if 'GEO-DIM 2G_3G_4G_5G' in wb.sheetnames:
            sheet2 = wb['GEO-DIM 2G_3G_4G_5G']
            rows2 = list(sheet2.iter_rows(values_only=True))
            if len(rows2) > 1:
                headers2 = [str(h).strip() if h else '' for h in rows2[0]]
                cells_to_upsert = []

                for r in rows2[1:]:
                    row_dict = dict(zip(headers2, r))
                    site_id = str(row_dict.get('Site ID') or '').strip()
                    cell_id = str(row_dict.get('Cell Id') or row_dict.get('LocalCellID') or '').strip()
                    if not site_id or not cell_id:
                        continue

                    cell_name = str(row_dict.get('CellName') or '').strip()
                    ne_name = str(row_dict.get('NE Name') or '').strip()
                    bts_id = str(row_dict.get('BTS ID/eNodeBID') or '').strip()
                    mcc = str(row_dict.get('MCC') or '619').strip()
                    mnc = str(row_dict.get('MNC') or '01').strip()
                    lac = str(row_dict.get('LAC') or row_dict.get('LAC ') or '').strip()
                    cgi = str(row_dict.get('CGI') or '').strip()
                    bsc = str(row_dict.get('BSC Name') or '').strip()
                    tech = str(row_dict.get('Technology') or '4G').strip()
                    lat = _dec(row_dict.get('Latitude'))
                    lng = _dec(row_dict.get('Longitude'))

                    cells_to_upsert.append({
                        'operator_code': 'orange',
                        'site_id': site_id,
                        'cell_id': cell_id,
                        'defaults': {
                            'bts_name': row_dict.get('BTS Name') or ne_name or site_id,
                            'cell_name': cell_name,
                            'ne_name': ne_name,
                            'bts_enodeb_id': bts_id,
                            'mcc': mcc,
                            'mnc': mnc,
                            'lac_tac': lac,
                            'cgi_ecgi': cgi,
                            'bsc_rnc_name': bsc,
                            'technology': tech,
                            'latitude': lat,
                            'longitude': lng,
                            'status': 'ACTIVE',
                        }
                    })

                with transaction.atomic():
                    for item in cells_to_upsert:
                        NetworkSectorCell.objects.update_or_create(
                            operator_code=item['operator_code'],
                            site_id=item['site_id'],
                            cell_id=item['cell_id'],
                            defaults=item['defaults'],
                        )
                        cells_created += 1

                self.stdout.write(f'Imported/updated {cells_created} sector cell records.')

        self.stdout.write(self.style.SUCCESS(
            f'Geo-Dimension import complete: {sites_created} physical sites + {cells_created} sector cells.'))
