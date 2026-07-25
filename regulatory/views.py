"""Regulatory Service — Views.

CRUD pattern mirrors ``reference/views.py``: HTML list + JSON API + POST
save + POST delete.  Engines live in ``regulatory/engines/``.  LEA views
are gated by :py:func:`regulatory.decorators.lawful_intercept_required`.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q
from django.http import JsonResponse, HttpResponse, Http404
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.views.decorators.http import require_POST

from .decorators import lawful_intercept_required
from .models import (
    RegulatoryProfile, RetailRevenue, RegulatoryReport,
    LeviedPeriod, LEARequest, LEAExtractionLog, QoSMetric,
    NetworkKPIDefinition, NetworkKPIEntry, NetworkKPIImportLog,
    DriveTestCampaign, DriveTestSample, DriveTestAnalysis,
    NetworkCellSite, NetworkCounterDefinition,
)


# =============================================================================
# Helpers
# =============================================================================

def _paginate(qs, page, per_page=25):
    total = qs.count()
    pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(int(page or 1), pages))
    offset = (page - 1) * per_page
    return qs[offset:offset + per_page], total, page, pages


def _decimal(value, default='0'):
    if value in (None, ''):
        return Decimal(default)
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)


def _int(value, default=0):
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def _date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _datetime(value):
    if not value:
        return None
    for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return timezone.make_aware(datetime.strptime(value, fmt))
        except (ValueError, TypeError):
            continue
    return None


# =============================================================================
# Index
# =============================================================================

@login_required
def index(request):
    return redirect('regulatory:report_list')


# =============================================================================
# 1. NATCOM Reports
# =============================================================================

@login_required
def report_list(request):
    return render(request, 'regulatory/natcom_reports.html', {
        'title': 'NATCOM Reports',
        'total': RegulatoryReport.objects.count(),
        'report_types': RegulatoryReport.ReportType.choices,
        'status_choices': RegulatoryReport.Status.choices,
    })


@login_required
def report_api(request):
    rtype = request.GET.get('report_type', '').strip()
    status = request.GET.get('status', '').strip()
    page = request.GET.get('page', 1)
    qs = RegulatoryReport.objects.all()
    if rtype:
        qs = qs.filter(report_type=rtype)
    if status:
        qs = qs.filter(status=status)
    rows, total, page, pages = _paginate(qs, page)
    data = [{
        'id': r.pk,
        'report_type': r.report_type,
        'report_type_label': r.get_report_type_display(),
        'period_start': r.period_start.isoformat(),
        'period_end': r.period_end.isoformat(),
        'status': r.status,
        'generated_at': r.generated_at.isoformat() if r.generated_at else '',
        'pdf_url': r.pdf_file.url if r.pdf_file else '',
        'xlsx_url': r.xlsx_file.url if r.xlsx_file else '',
    } for r in rows]
    return JsonResponse({'records': data, 'total': total, 'page': page, 'pages': pages})


@login_required
@require_POST
def report_generate(request):
    rtype = request.POST.get('report_type', RegulatoryReport.ReportType.TRAFFIC)
    start = _date(request.POST.get('period_start'))
    end = _date(request.POST.get('period_end'))
    if not start or not end:
        return JsonResponse({'success': False, 'error': 'period_start + period_end required'})
    try:
        from core.tasks import enqueue_job
        from .tasks import task_generate_report
        job = enqueue_job(
            task=task_generate_report,
            job_type='regulatory.generate_report',
            label=f'NATCOM {rtype} report for {start}..{end}',
            user=request.user,
            params={'report_type': rtype, 'period_start': start.isoformat(),
                    'period_end': end.isoformat()},
            args=(rtype, start.isoformat(), end.isoformat(),
                   request.user.pk if request.user.is_authenticated else None),
        )
        return JsonResponse({
            'success': True, 'job_id': job.pk,
            'job_url': f'/jobs/{job.pk}/',
            'message': 'Report generation queued.',
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def report_pdf(request, pk):
    r = get_object_or_404(RegulatoryReport, pk=pk)
    if r.pdf_file:
        try:
            data = r.pdf_file.read()
            r.pdf_file.close()
            resp = HttpResponse(data, content_type='application/pdf')
            resp['Content-Disposition'] = f'inline; filename="natcom_{r.report_type}_{r.period_start}_{r.period_end}.pdf"'
            return resp
        except Exception:
            pass
    raise Http404('PDF not available; regenerate the report.')


@login_required
def report_xlsx(request, pk):
    r = get_object_or_404(RegulatoryReport, pk=pk)
    if r.xlsx_file:
        try:
            data = r.xlsx_file.read()
            r.xlsx_file.close()
            resp = HttpResponse(
                data,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            resp['Content-Disposition'] = f'attachment; filename="natcom_{r.report_type}_{r.period_start}_{r.period_end}.xlsx"'
            return resp
        except Exception:
            pass
    raise Http404('Excel not available; regenerate the report.')


@login_required
@require_POST
def report_delete(request, pk):
    get_object_or_404(RegulatoryReport, pk=pk).delete()
    return JsonResponse({'success': True})


@login_required
@require_POST
def report_set_status(request, pk):
    r = get_object_or_404(RegulatoryReport, pk=pk)
    new_status = request.POST.get('status')
    if new_status not in dict(RegulatoryReport.Status.choices):
        return JsonResponse({'success': False, 'error': 'Invalid status'})
    r.status = new_status
    if new_status == RegulatoryReport.Status.SUBMITTED and not r.submitted_at:
        r.submitted_at = timezone.now()
    r.save()
    return JsonResponse({'success': True, 'status': r.status})


# =============================================================================
# 2. Levy & USF
# =============================================================================

@login_required
def levy_list(request):
    return render(request, 'regulatory/levy.html', {
        'title': 'Levy & USF',
        'total': LeviedPeriod.objects.count(),
        'profile': RegulatoryProfile.get_or_create_default(),
        'status_choices': LeviedPeriod.Status.choices,
    })


@login_required
def levy_api(request):
    page = request.GET.get('page', 1)
    qs = LeviedPeriod.objects.all()
    rows, total, page, pages = _paginate(qs, page)
    data = [{
        'id': r.pk,
        'period_year': r.period_year,
        'period_month': r.period_month,
        'interconnect_inbound': str(r.interconnect_inbound),
        'interconnect_outbound': str(r.interconnect_outbound),
        'retail_total': str(r.retail_total),
        'gross_revenue': str(r.gross_revenue),
        'levy_pct': str(r.levy_pct),
        'usf_pct': str(r.usf_pct),
        'levy_amount': str(r.levy_amount),
        'usf_amount': str(r.usf_amount),
        'total_payable': str(r.total_payable),
        'currency': r.currency,
        'status': r.status,
        'due_date': r.due_date.isoformat() if r.due_date else '',
        'paid_at': r.paid_at.isoformat() if r.paid_at else '',
        'payment_reference': r.payment_reference,
    } for r in rows]
    return JsonResponse({'records': data, 'total': total, 'page': page, 'pages': pages})


@login_required
@require_POST
def levy_compute(request):
    year = _int(request.POST.get('period_year'))
    month = _int(request.POST.get('period_month'))
    if not (year and 1 <= month <= 12):
        return JsonResponse({'success': False, 'error': 'period_year + period_month required'})
    try:
        from .engines.levy import compute_levy
        levy = compute_levy(year, month, user=request.user if request.user.is_authenticated else None)
        return JsonResponse({
            'success': True, 'id': levy.pk,
            'gross_revenue': str(levy.gross_revenue),
            'levy_amount': str(levy.levy_amount),
            'usf_amount': str(levy.usf_amount),
            'total_payable': str(levy.total_payable),
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def levy_mark_paid(request, pk):
    levy = get_object_or_404(LeviedPeriod, pk=pk)
    try:
        from .engines.levy import mark_levy_paid
        mark_levy_paid(
            levy,
            payment_date=_date(request.POST.get('payment_date')) or date.today(),
            reference=request.POST.get('payment_reference', '').strip(),
            user=request.user if request.user.is_authenticated else None,
        )
        return JsonResponse({'success': True, 'status': levy.status})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def levy_delete(request, pk):
    get_object_or_404(LeviedPeriod, pk=pk).delete()
    return JsonResponse({'success': True})


# =============================================================================
# 3. Retail Revenue (manual entry)
# =============================================================================

@login_required
def retail_list(request):
    return render(request, 'regulatory/retail_revenue.html', {
        'title': 'Retail Revenue',
        'total': RetailRevenue.objects.count(),
    })


@login_required
def retail_api(request):
    q = request.GET.get('q', '').strip()
    page = request.GET.get('page', 1)
    qs = RetailRevenue.objects.all()
    if q:
        qs = qs.filter(notes__icontains=q)
    rows, total, page, pages = _paginate(qs, page)
    data = [{
        'id': r.pk,
        'period_year': r.period_year,
        'period_month': r.period_month,
        'voice_revenue': str(r.voice_revenue),
        'sms_revenue': str(r.sms_revenue),
        'data_revenue': str(r.data_revenue),
        'other_revenue': str(r.other_revenue),
        'total': str(r.total),
        'currency': r.currency,
        'notes': r.notes,
    } for r in rows]
    return JsonResponse({'records': data, 'total': total, 'page': page, 'pages': pages})


@login_required
@require_POST
def retail_save(request):
    pk = request.POST.get('id')
    try:
        if pk:
            obj = RetailRevenue.objects.get(pk=pk)
        else:
            obj = RetailRevenue()
            obj.created_by = request.user if request.user.is_authenticated else None
        obj.period_year = _int(request.POST.get('period_year'))
        obj.period_month = _int(request.POST.get('period_month'))
        obj.voice_revenue = _decimal(request.POST.get('voice_revenue'))
        obj.sms_revenue = _decimal(request.POST.get('sms_revenue'))
        obj.data_revenue = _decimal(request.POST.get('data_revenue'))
        obj.other_revenue = _decimal(request.POST.get('other_revenue'))
        obj.currency = request.POST.get('currency', 'SLE').strip().upper()
        obj.notes = request.POST.get('notes', '').strip()
        obj.save()
        return JsonResponse({'success': True, 'id': obj.pk, 'total': str(obj.total)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def retail_delete(request, pk):
    get_object_or_404(RetailRevenue, pk=pk).delete()
    return JsonResponse({'success': True})


# =============================================================================
# 4. Lawful Intercept (gated)
# =============================================================================

@lawful_intercept_required
def intercept_list(request):
    return render(request, 'regulatory/intercept.html', {
        'title': 'Lawful Intercept',
        'total': LEARequest.objects.count(),
        'status_choices': LEARequest.Status.choices,
    })


@lawful_intercept_required
def intercept_api(request):
    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()
    page = request.GET.get('page', 1)
    qs = LEARequest.objects.all()
    if status:
        qs = qs.filter(status=status)
    if q:
        qs = qs.filter(
            Q(case_number__icontains=q) | Q(requesting_agency__icontains=q) |
            Q(officer_name__icontains=q) | Q(filter_msisdn__icontains=q) |
            Q(filter_imsi__icontains=q) | Q(filter_imei__icontains=q)
        )
    rows, total, page, pages = _paginate(qs, page)
    data = [{
        'id': r.pk,
        'case_number': r.case_number,
        'requesting_agency': r.requesting_agency,
        'officer_name': r.officer_name,
        'officer_contact': r.officer_contact,
        'legal_basis': r.legal_basis,
        'filter_msisdn': r.filter_msisdn,
        'filter_imsi': r.filter_imsi,
        'filter_imei': r.filter_imei,
        'filter_cell_id': r.filter_cell_id,
        'filter_start': r.filter_start.isoformat() if r.filter_start else '',
        'filter_end': r.filter_end.isoformat() if r.filter_end else '',
        'status': r.status,
        'opened_at': r.opened_at.isoformat() if r.opened_at else '',
        'fulfilled_at': r.fulfilled_at.isoformat() if r.fulfilled_at else '',
        'extractions': r.extractions.count(),
        'notes': r.notes,
    } for r in rows]
    return JsonResponse({'records': data, 'total': total, 'page': page, 'pages': pages})


@lawful_intercept_required
@require_POST
def intercept_save(request):
    pk = request.POST.get('id')
    try:
        if pk:
            obj = LEARequest.objects.get(pk=pk)
        else:
            obj = LEARequest()
            obj.opened_by = request.user if request.user.is_authenticated else None
        obj.case_number = request.POST.get('case_number', '').strip()
        obj.requesting_agency = request.POST.get('requesting_agency', '').strip()
        obj.officer_name = request.POST.get('officer_name', '').strip()
        obj.officer_contact = request.POST.get('officer_contact', '').strip()
        obj.legal_basis = request.POST.get('legal_basis', '').strip()
        obj.filter_msisdn = request.POST.get('filter_msisdn', '').strip()
        obj.filter_imsi = request.POST.get('filter_imsi', '').strip()
        obj.filter_imei = request.POST.get('filter_imei', '').strip()
        obj.filter_cell_id = request.POST.get('filter_cell_id', '').strip()
        obj.filter_start = _datetime(request.POST.get('filter_start')) or obj.filter_start
        obj.filter_end = _datetime(request.POST.get('filter_end')) or obj.filter_end
        obj.status = request.POST.get('status', LEARequest.Status.OPEN)
        obj.notes = request.POST.get('notes', '').strip()
        obj.save()
        return JsonResponse({'success': True, 'id': obj.pk})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@lawful_intercept_required
def intercept_detail(request, pk):
    req = get_object_or_404(LEARequest.objects.prefetch_related('extractions'), pk=pk)
    return render(request, 'regulatory/intercept_detail.html', {
        'title': f'LEA Request {req.case_number}',
        'request_obj': req,
        'status_choices': LEARequest.Status.choices,
    })


@lawful_intercept_required
@require_POST
def intercept_execute(request, pk):
    req = get_object_or_404(LEARequest, pk=pk)
    try:
        from .engines.intercept import run_lea_query
        limit = _int(request.POST.get('limit'), 100)
        rows = run_lea_query(req, limit=limit)
        return JsonResponse({'success': True, 'rows': rows, 'count': len(rows)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@lawful_intercept_required
@require_POST
def intercept_export(request, pk):
    req = get_object_or_404(LEARequest, pk=pk)
    try:
        from core.tasks import enqueue_job
        from .tasks import task_intercept_export
        job = enqueue_job(
            task=task_intercept_export,
            job_type='regulatory.intercept_export',
            label=f'LEA evidentiary export: {req.case_number}',
            user=request.user,
            params={'case_number': req.case_number,
                    'filter_msisdn': req.filter_msisdn,
                    'filter_imsi': req.filter_imsi,
                    'filter_imei': req.filter_imei,
                    'filter_start': req.filter_start.isoformat() if req.filter_start else '',
                    'filter_end':   req.filter_end.isoformat()   if req.filter_end   else ''},
            args=(req.pk, request.user.pk if request.user.is_authenticated else None),
        )
        return JsonResponse({'success': True, 'job_id': job.pk,
                              'job_url': f'/jobs/{job.pk}/',
                              'message': 'Export queued.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@lawful_intercept_required
def intercept_download_extraction(request, pk):
    ext = get_object_or_404(LEAExtractionLog, pk=pk)
    if not ext.export_file:
        raise Http404('Export file missing')
    try:
        data = ext.export_file.read()
        ext.export_file.close()
    except Exception:
        raise Http404('Export file unreadable')
    resp = HttpResponse(data, content_type='text/csv')
    resp['Content-Disposition'] = (
        f'attachment; filename="LEA_{ext.request.case_number}_{ext.pk}.csv"'
    )
    return resp


@lawful_intercept_required
@require_POST
def intercept_delete(request, pk):
    get_object_or_404(LEARequest, pk=pk).delete()
    return JsonResponse({'success': True})


# =============================================================================
# 5. QoS / KPIs
# =============================================================================

@login_required
def qos_view(request):
    return render(request, 'regulatory/qos.html', {
        'title': 'QoS / KPIs',
        'total': QoSMetric.objects.count(),
    })


@login_required
def qos_api(request):
    granularity = request.GET.get('granularity', 'DAILY')
    start = _date(request.GET.get('start'))
    end = _date(request.GET.get('end'))
    qs = QoSMetric.objects.filter(granularity=granularity).order_by('metric_date')
    if start:
        qs = qs.filter(metric_date__gte=start)
    if end:
        qs = qs.filter(metric_date__lte=end)
    data = [{
        'date': r.metric_date.isoformat(),
        'total_calls': r.total_calls,
        'successful_calls': r.successful_calls,
        'dropped_calls': r.dropped_calls,
        'failed_calls': r.failed_calls,
        'asr_pct': float(r.asr_pct),
        'acd_seconds': float(r.acd_seconds),
        'drop_rate_pct': float(r.drop_rate_pct),
        'availability_pct': float(r.availability_pct),
        'source': r.source,
    } for r in qs]
    return JsonResponse({'success': True, 'data': data})


@login_required
@require_POST
def qos_refresh(request):
    try:
        from .engines.qos import compute_daily_qos
        target_date = _date(request.POST.get('date')) or date.today()
        metric = compute_daily_qos(target_date)
        return JsonResponse({'success': True, 'date': metric.metric_date.isoformat(),
                              'asr_pct': float(metric.asr_pct),
                              'drop_rate_pct': float(metric.drop_rate_pct)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# =============================================================================
# Regulatory Profile (admin)
# =============================================================================

def _admin_required(u):
    return u.is_authenticated and u.is_superuser


@user_passes_test(_admin_required)
@require_POST
def profile_save(request):
    p = RegulatoryProfile.get_or_create_default()
    p.regulator_name = request.POST.get('regulator_name', p.regulator_name).strip() or 'NATCOM'
    p.contact_email = request.POST.get('contact_email', p.contact_email).strip()
    p.address = request.POST.get('address', p.address).strip()
    p.phone = request.POST.get('phone', p.phone).strip()
    p.levy_pct = _decimal(request.POST.get('levy_pct'), str(p.levy_pct))
    p.usf_pct = _decimal(request.POST.get('usf_pct'), str(p.usf_pct))
    p.home_currency = request.POST.get('home_currency', p.home_currency).strip() or 'SLE'
    p.updated_by = request.user
    p.save()
    return JsonResponse({'success': True})


# =============================================================================
# 6. Network Performance Monitoring (PM KPIs)
# =============================================================================

@login_required
def network_performance_view(request):
    return render(request, 'regulatory/network_performance.html', {
        'title': 'Network Performance (PM KPIs)',
        'kpi_defs': NetworkKPIDefinition.objects.filter(is_active=True),
        'total_entries': NetworkKPIEntry.objects.count(),
    })


@login_required
def network_performance_api(request):
    start = _date(request.GET.get('start'))
    end = _date(request.GET.get('end'))
    code = request.GET.get('code', '').strip().upper()
    operator = request.GET.get('operator', '').strip().lower()
    region = request.GET.get('region', '').strip()
    district = request.GET.get('district', '').strip()
    page = request.GET.get('page', 1)

    qs = NetworkKPIEntry.objects.select_related('kpi').all()
    if start:
        qs = qs.filter(period_date__gte=start)
    if end:
        qs = qs.filter(period_date__lte=end)
    if code:
        qs = qs.filter(kpi__code=code)
    if operator:
        qs = qs.filter(operator_code=operator)
    if region:
        qs = qs.filter(region__icontains=region)
    if district:
        qs = qs.filter(district__icontains=district)

    rows, total, page, pages = _paginate(qs, page)
    data = [{
        'id': r.pk,
        'kpi_code': r.kpi.code,
        'kpi_name': r.kpi.name,
        'unit': r.kpi.unit,
        'period_date': r.period_date.isoformat(),
        'granularity': r.granularity,
        'operator_code': r.operator_code,
        'region': r.region,
        'district': r.district,
        'cell_id': r.cell_id,
        'value': str(r.value),
        'natca_threshold': str(r.kpi.natca_threshold),
        'is_compliant': r.is_compliant,
        'source': r.source,
        'notes': r.notes,
    } for r in rows]
    return JsonResponse({'records': data, 'total': total, 'page': page, 'pages': pages})


@login_required
def network_performance_comparison_api(request):
    """Return multi-operator comparative analysis grid."""
    start = _date(request.GET.get('start'))
    end = _date(request.GET.get('end'))
    region = request.GET.get('region', '').strip()
    district = request.GET.get('district', '').strip()

    from .engines.network_kpi import get_operator_comparison_matrix
    matrix = get_operator_comparison_matrix(start_date=start, end_date=end, region=region, district=district)
    return JsonResponse({'success': True, 'matrix': matrix})


@login_required
@require_POST
def network_performance_import(request):
    """Handle bulk file import (CSV, ZIP, TAR.GZ)."""
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'No file uploaded'})
    uploaded_file = request.FILES['file']
    try:
        from .engines.network_kpi import import_kpi_file
        res = import_kpi_file(
            file_obj=uploaded_file,
            filename=uploaded_file.name,
            channel='CSV_IMPORT',
            user=request.user,
        )
        return JsonResponse(res)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_POST
def network_performance_api_push(request):
    """REST API endpoint for automated NMS/OSS push of KPI payload."""
    import json
    try:
        data = json.loads(request.body.decode('utf-8'))
        rows = data if isinstance(data, list) else data.get('records', [])
        from .engines.network_kpi import process_kpi_rows
        res = process_kpi_rows(
            rows=rows,
            filename=f"API_Push_{timezone.now():%Y%m%d_%H%M%S}.json",
            channel='API',
            user=request.user if request.user.is_authenticated else None,
        )
        return JsonResponse(res)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def network_performance_save(request):
    pk = request.POST.get('id')
    code = request.POST.get('kpi_code', '').strip().upper()
    try:
        kpi_def = NetworkKPIDefinition.objects.get(code=code)
        if pk:
            obj = NetworkKPIEntry.objects.get(pk=pk)
        else:
            obj = NetworkKPIEntry()

        obj.kpi = kpi_def
        obj.period_date = _date(request.POST.get('period_date')) or date.today()
        obj.granularity = request.POST.get('granularity', 'DAILY').upper()
        obj.operator_code = request.POST.get('operator_code', 'orange').strip().lower() or 'orange'
        obj.region = request.POST.get('region', 'NATIONAL').strip() or 'NATIONAL'
        obj.district = request.POST.get('district', '').strip()
        obj.cell_id = request.POST.get('cell_id', '').strip()
        obj.value = _decimal(request.POST.get('value'))
        from .engines.network_kpi import check_kpi_compliance
        obj.is_compliant = check_kpi_compliance(kpi_def, obj.value)
        obj.source = 'MANUAL'
        obj.notes = request.POST.get('notes', '').strip()
        obj.save()

        # Recompute QoS score if needed
        from .engines.network_kpi import compute_qos_compliance_score
        compute_qos_compliance_score(obj.period_date, obj.operator_code, obj.region, obj.district)

        return JsonResponse({'success': True, 'id': obj.pk, 'is_compliant': obj.is_compliant})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def network_performance_delete(request, pk):
    get_object_or_404(NetworkKPIEntry, pk=pk).delete()
    return JsonResponse({'success': True})


# =============================================================================
# 7. Drive Test Management
# =============================================================================

@login_required
def drive_test_list(request):
    return render(request, 'regulatory/drive_test.html', {
        'title': 'Drive Test Campaigns',
        'total': DriveTestCampaign.objects.count(),
        'status_choices': DriveTestCampaign.Status.choices,
    })


@login_required
def drive_test_api(request):
    status = request.GET.get('status', '').strip()
    q = request.GET.get('q', '').strip()
    page = request.GET.get('page', 1)

    qs = DriveTestCampaign.objects.select_related('analysis').all()
    if status:
        qs = qs.filter(status=status)
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(region__icontains=q) | Q(route_description__icontains=q))

    rows, total, page, pages = _paginate(qs, page)
    data = []
    for r in rows:
        analysis = getattr(r, 'analysis', None)
        data.append({
            'id': r.pk,
            'name': r.name,
            'test_date': r.test_date.isoformat(),
            'region': r.region,
            'technology': r.technology,
            'tool_used': r.tool_used,
            'operator_name': r.operator_name,
            'status': r.status,
            'raw_file_url': r.raw_file.url if r.raw_file else '',
            'total_samples': analysis.total_samples if analysis else 0,
            'coverage_pct': str(analysis.coverage_pct) if analysis else '0.00',
            'avg_rsrp': str(analysis.avg_rsrp) if analysis else '0.00',
            'avg_dl_tp': str(analysis.avg_dl_throughput) if analysis else '0.00',
            'natca_compliant': analysis.natca_compliant if analysis else False,
        })
    return JsonResponse({'records': data, 'total': total, 'page': page, 'pages': pages})


@login_required
@require_POST
def drive_test_upload(request):
    name = request.POST.get('name', '').strip()
    test_date = _date(request.POST.get('test_date')) or date.today()
    region = request.POST.get('region', 'WESTERN_AREA').strip()
    tech = request.POST.get('technology', '4G').strip()
    tool = request.POST.get('tool_used', 'TEMS').strip()

    if not name or 'file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'Name and measurement file required'})

    f = request.FILES['file']
    fn_lower = f.name.lower()
    fmt = 'csv'
    for ext in ('zip', 'tar.gz', 'tgz', 'trp', 'lpg', 'nmf', 'csv'):
        if fn_lower.endswith('.' + ext):
            fmt = ext
            break

    campaign = DriveTestCampaign.objects.create(
        name=name,
        test_date=test_date,
        region=region,
        technology=tech,
        tool_used=tool,
        raw_file=f,
        file_format=fmt,
        created_by=request.user,
    )

    try:
        from .engines.drive_test import parse_drive_test_file, analyse_campaign
        f.seek(0)
        sample_count = parse_drive_test_file(f, f.name, campaign)
        analyse_campaign(campaign, user=request.user)
        return JsonResponse({'success': True, 'id': campaign.pk, 'samples': sample_count})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def drive_test_detail(request, pk):
    campaign = get_object_or_404(DriveTestCampaign.objects.select_related('analysis'), pk=pk)
    analysis = getattr(campaign, 'analysis', None)
    return render(request, 'regulatory/drive_test_detail.html', {
        'title': f'Drive Test: {campaign.name}',
        'campaign': campaign,
        'analysis': analysis,
    })


@login_required
def drive_test_samples_api(request, pk):
    campaign = get_object_or_404(DriveTestCampaign, pk=pk)
    samples = DriveTestSample.objects.filter(campaign=campaign)[:2000]
    data = [{
        'id': s.pk,
        'lat': float(s.latitude),
        'lng': float(s.longitude),
        'rsrp': float(s.rsrp) if s.rsrp is not None else None,
        'rsrq': float(s.rsrq) if s.rsrq is not None else None,
        'sinr': float(s.sinr) if s.sinr is not None else None,
        'dl_tp': float(s.dl_throughput) if s.dl_throughput is not None else None,
        'ul_tp': float(s.ul_throughput) if s.ul_throughput is not None else None,
        'mos': float(s.voice_mos) if s.voice_mos is not None else None,
        'cell_id': s.cell_id,
    } for s in samples]
    return JsonResponse({'success': True, 'samples': data})


@login_required
@require_POST
def drive_test_analyse(request, pk):
    campaign = get_object_or_404(DriveTestCampaign, pk=pk)
    try:
        from .engines.drive_test import analyse_campaign
        analysis = analyse_campaign(campaign, user=request.user)
        return JsonResponse({'success': True, 'compliant': analysis.natca_compliant})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def drive_test_delete(request, pk):
    get_object_or_404(DriveTestCampaign, pk=pk).delete()
    return JsonResponse({'success': True})


# =============================================================================
# 8. Cell Site & Geo Dimension (GeoDim) Management
# =============================================================================

@login_required
def site_list_view(request):
    return render(request, 'regulatory/sites.html', {
        'title': 'Cell Site & Geo Inventory',
        'total': NetworkCellSite.objects.count(),
        'status_choices': NetworkCellSite.Status.choices,
    })


@login_required
def site_api(request):
    operator = request.GET.get('operator', '').strip().lower()
    region = request.GET.get('region', '').strip()
    district = request.GET.get('district', '').strip()
    tech = request.GET.get('technology', '').strip()
    status = request.GET.get('status', '').strip()
    mode = request.GET.get('mode', '').strip().lower()  # 'physical' or 'sectors'
    q = request.GET.get('q', '').strip()
    page = request.GET.get('page', 1)

    qs = NetworkCellSite.objects.all()
    if mode == 'physical':
        qs = qs.filter(cell_id='')
    elif mode == 'sectors':
        qs = qs.exclude(cell_id='')

    if operator:
        qs = qs.filter(operator_code=operator)
    if region:
        qs = qs.filter(region__icontains=region)
    if district:
        qs = qs.filter(district__icontains=district)
    if tech:
        qs = qs.filter(technology__icontains=tech)
    if status:
        qs = qs.filter(status=status)
    if q:
        qs = qs.filter(
            Q(site_id__icontains=q) | Q(site_name__icontains=q) |
            Q(cell_id__icontains=q) | Q(cell_name__icontains=q) |
            Q(cgi_ecgi__icontains=q) | Q(bsc_rnc_name__icontains=q) |
            Q(location__icontains=q) | Q(chiefdom__icontains=q)
        )

    rows, total, page, pages = _paginate(qs, page)
    data = [{
        'id': r.pk,
        'operator_code': r.operator_code,
        'site_id': r.site_id,
        'site_name': r.site_name,
        'cell_id': r.cell_id,
        'cell_name': r.cell_name,
        'ne_name': r.ne_name,
        'bts_enodeb_id': r.bts_enodeb_id,
        'mcc': r.mcc,
        'mnc': r.mnc,
        'lac_tac': r.lac_tac,
        'cgi_ecgi': r.cgi_ecgi,
        'bsc_rnc_name': r.bsc_rnc_name,
        'technology': r.technology,
        'classification': r.classification,
        'natca_classification': r.natca_classification,
        'site_owner': r.site_owner,
        'region': r.region,
        'district': r.district,
        'chiefdom': r.chiefdom,
        'location': r.location,
        'latitude': float(r.latitude) if r.latitude is not None else None,
        'longitude': float(r.longitude) if r.longitude is not None else None,
        'height_m': float(r.height_m) if r.height_m is not None else None,
        'azimuth': r.azimuth,
        'on_air_date': r.on_air_date.isoformat() if r.on_air_date else None,
        'site_type': r.site_type,
        'status': r.status,
        'notes': r.notes,
    } for r in rows]
    return JsonResponse({'records': data, 'total': total, 'page': page, 'pages': pages})


@login_required
@require_POST
def site_save(request):
    pk = request.POST.get('id')
    try:
        if pk:
            obj = NetworkCellSite.objects.get(pk=pk)
        else:
            obj = NetworkCellSite()

        obj.operator_code = request.POST.get('operator_code', 'orange').strip().lower() or 'orange'
        obj.site_id = request.POST.get('site_id', '').strip()
        obj.site_name = request.POST.get('site_name', '').strip()
        obj.cell_id = request.POST.get('cell_id', '').strip()
        obj.technology = request.POST.get('technology', '4G').strip()
        obj.region = request.POST.get('region', 'WESTERN_AREA').strip()
        obj.district = request.POST.get('district', '').strip()
        obj.chiefdom_town = request.POST.get('chiefdom_town', '').strip()

        lat = request.POST.get('latitude')
        lng = request.POST.get('longitude')
        h = request.POST.get('height_m')
        az = request.POST.get('azimuth')

        obj.latitude = _decimal(lat, None) if lat else None
        obj.longitude = _decimal(lng, None) if lng else None
        obj.height_m = _decimal(h, None) if h else None
        obj.azimuth = _int(az, None) if az else None

        obj.status = request.POST.get('status', NetworkCellSite.Status.ACTIVE)
        obj.notes = request.POST.get('notes', '').strip()
        obj.save()

        return JsonResponse({'success': True, 'id': obj.pk})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def site_import(request):
    """Handle bulk Excel (.xlsx) and CSV (.csv) import of Cell Site & Geo Inventory."""
    import csv, io
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'No file uploaded'})

    uploaded_file = request.FILES['file']
    fn = uploaded_file.name.lower()
    count = 0
    errors = []

    try:
        if fn.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            for sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                rows = list(sheet.iter_rows(values_only=True))
                if len(rows) <= 1:
                    continue
                headers = [str(h).strip() if h else '' for h in rows[0]]
                for idx, r in enumerate(rows[1:], start=2):
                    row_dict = dict(zip(headers, r))
                    s_id = str(row_dict.get('SITE ID') or row_dict.get('Site ID') or row_dict.get('site_id') or '').strip()
                    c_id = str(row_dict.get('Cell Id') or row_dict.get('cell_id') or row_dict.get('LocalCellID') or '').strip()
                    if not s_id:
                        continue

                    s_name = str(row_dict.get('SITE NAME') or row_dict.get('NE Name') or row_dict.get('site_name') or s_id).strip()
                    op = str(row_dict.get('operator_code') or row_dict.get('operator') or 'orange').strip().lower()

                    NetworkCellSite.objects.update_or_create(
                        operator_code=op,
                        site_id=s_id,
                        cell_id=c_id,
                        defaults={
                            'site_name': s_name,
                            'cell_name': str(row_dict.get('CellName') or row_dict.get('cell_name') or '').strip(),
                            'ne_name': str(row_dict.get('NE Name') or row_dict.get('ne_name') or '').strip(),
                            'bts_enodeb_id': str(row_dict.get('BTS ID/eNodeBID') or row_dict.get('bts_enodeb_id') or '').strip(),
                            'mcc': str(row_dict.get('MCC') or '619').strip(),
                            'mnc': str(row_dict.get('MNC') or '01').strip(),
                            'lac_tac': str(row_dict.get('LAC') or row_dict.get('LAC ') or row_dict.get('lac_tac') or '').strip(),
                            'cgi_ecgi': str(row_dict.get('CGI') or row_dict.get('cgi_ecgi') or '').strip(),
                            'bsc_rnc_name': str(row_dict.get('BSC Name') or row_dict.get('bsc_rnc_name') or '').strip(),
                            'technology': str(row_dict.get('Technology') or row_dict.get('technology') or '4G').strip(),
                            'classification': str(row_dict.get('Classification') or row_dict.get('classification') or '').strip(),
                            'natca_classification': str(row_dict.get('NAtCa Sites Classification') or row_dict.get('natca_classification') or '').strip(),
                            'site_owner': str(row_dict.get('OWNER') or row_dict.get('site_owner') or '').strip(),
                            'region': str(row_dict.get('Region') or row_dict.get('region') or 'Western Area').strip(),
                            'district': str(row_dict.get('District') or row_dict.get('district') or '').strip(),
                            'chiefdom': str(row_dict.get('Chiefdom') or row_dict.get('chiefdom') or '').strip(),
                            'location': str(row_dict.get('Location') or row_dict.get('location') or '').strip(),
                            'latitude': _decimal(row_dict.get('LATITUDE') or row_dict.get('Latitude') or row_dict.get('latitude'), None),
                            'longitude': _decimal(row_dict.get('LONGITUDE') or row_dict.get('Longitude') or row_dict.get('longitude'), None),
                            'height_m': _decimal(row_dict.get('Tower Height') or row_dict.get('height_m'), None),
                            'site_type': str(row_dict.get('Site Type') or row_dict.get('site_type') or '').strip(),
                            'status': NetworkCellSite.Status.ACTIVE,
                        }
                    )
                    count += 1
        else:
            # CSV file
            content = uploaded_file.read().decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(io.StringIO(content))
            for idx, r in enumerate(reader, start=1):
                s_id = str(r.get('site_id') or r.get('SITE ID') or r.get('Site ID') or '').strip()
                c_id = str(r.get('cell_id') or r.get('Cell Id') or '').strip()
                if not s_id:
                    continue

                op = str(r.get('operator_code') or r.get('operator') or 'orange').strip().lower()
                NetworkCellSite.objects.update_or_create(
                    operator_code=op,
                    site_id=s_id,
                    cell_id=c_id,
                    defaults={
                        'site_name': str(r.get('site_name') or r.get('SITE NAME') or s_id).strip(),
                        'cell_name': str(r.get('cell_name') or r.get('CellName') or '').strip(),
                        'technology': str(r.get('technology') or r.get('Technology') or '4G').strip(),
                        'region': str(r.get('region') or r.get('Region') or 'Western Area').strip(),
                        'district': str(r.get('district') or r.get('District') or '').strip(),
                        'latitude': _decimal(r.get('latitude') or r.get('LATITUDE') or r.get('Latitude'), None),
                        'longitude': _decimal(r.get('longitude') or r.get('LONGITUDE') or r.get('Longitude'), None),
                        'status': NetworkCellSite.Status.ACTIVE,
                    }
                )
                count += 1

        return JsonResponse({'success': True, 'count': count, 'error_count': len(errors)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def site_delete(request, pk):
    get_object_or_404(NetworkCellSite, pk=pk).delete()
    return JsonResponse({'success': True})


@login_required
def site_download_template(request):
    """Download sample CSV template for Cell Site Inventory import."""
    import csv, io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'operator_code', 'site_id', 'site_name', 'cell_id', 'technology',
        'region', 'district', 'chiefdom_town', 'latitude', 'longitude',
        'height_m', 'azimuth', 'status', 'notes'
    ])
    writer.writerow([
        'orange', 'FTW001', 'Lumley Beach Tower', 'FTW001_1', '4G',
        'WESTERN_AREA', 'Western Area Urban', 'Lumley', '8.4842000', '-13.2301000',
        '45.0', '120', 'ACTIVE', 'Primary Lumley Sector A'
    ])
    writer.writerow([
        'africell', 'AF_BO01', 'Bo City Central Tower', 'AF_BO01_1', '4G',
        'SOUTHERN', 'Bo', 'Bo Town', '7.9647000', '-11.7383000',
        '50.0', '0', 'ACTIVE', 'Bo Main Switch Hub'
    ])

    resp = HttpResponse(output.getvalue(), content_type='text/csv')
    resp['Content-Disposition'] = 'attachment; filename="cell_sites_template.csv"'
    return resp


# =============================================================================
# 9. Counter Dictionary / Inventory Catalog
# =============================================================================

@login_required
def counter_list_view(request):
    return render(request, 'regulatory/counters.html', {
        'title': 'PM Counter Dictionary',
        'total': NetworkCounterDefinition.objects.count(),
        'kpi_defs': NetworkKPIDefinition.objects.filter(is_active=True),
        'role_choices': NetworkCounterDefinition.FormulaRole.choices,
    })


@login_required
def counter_api(request):
    vendor = request.GET.get('vendor', '').strip()
    ne = request.GET.get('network_element', '').strip()
    tech = request.GET.get('technology', '').strip()
    kpi_code = request.GET.get('kpi_code', '').strip().upper()
    q = request.GET.get('q', '').strip()
    page = request.GET.get('page', 1)

    qs = NetworkCounterDefinition.objects.all()
    if vendor:
        qs = qs.filter(vendor__icontains=vendor)
    if ne:
        qs = qs.filter(network_element__icontains=ne)
    if tech:
        qs = qs.filter(technology=tech)
    if kpi_code:
        qs = qs.filter(kpi_code=kpi_code)
    if q:
        qs = qs.filter(
            Q(counter_id__icontains=q) | Q(counter_name__icontains=q) | Q(description__icontains=q)
        )

    rows, total, page, pages = _paginate(qs, page)
    data = [{
        'id': r.pk,
        'counter_id': r.counter_id,
        'counter_name': r.counter_name,
        'vendor': r.vendor,
        'network_element': r.network_element,
        'technology': r.technology,
        'kpi_code': r.kpi_code,
        'formula_role': r.formula_role,
        'description': r.description,
        'is_active': r.is_active,
    } for r in rows]
    return JsonResponse({'records': data, 'total': total, 'page': page, 'pages': pages})


@login_required
@require_POST
def counter_save(request):
    pk = request.POST.get('id')
    try:
        if pk:
            obj = NetworkCounterDefinition.objects.get(pk=pk)
        else:
            obj = NetworkCounterDefinition()

        obj.counter_id = request.POST.get('counter_id', '').strip()
        obj.counter_name = request.POST.get('counter_name', '').strip()
        obj.vendor = request.POST.get('vendor', 'Huawei').strip() or 'Huawei'
        obj.network_element = request.POST.get('network_element', 'eNodeB').strip() or 'eNodeB'
        obj.technology = request.POST.get('technology', '4G').strip()
        obj.kpi_code = request.POST.get('kpi_code', '').strip().upper()
        obj.formula_role = request.POST.get('formula_role', NetworkCounterDefinition.FormulaRole.NUMERATOR)
        obj.description = request.POST.get('description', '').strip()
        obj.is_active = request.POST.get('is_active', 'true').lower() in ('true', '1', 'yes')
        obj.save()

        return JsonResponse({'success': True, 'id': obj.pk})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def counter_import(request):
    """Bulk CSV import for Counter Dictionary catalog."""
    import csv, io
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'No file uploaded'})

    uploaded_file = request.FILES['file']
    try:
        content = uploaded_file.read().decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(io.StringIO(content))
        count = 0
        errors = []

        for idx, r in enumerate(reader, start=1):
            c_id = str(r.get('counter_id') or r.get('id') or '').strip()
            c_name = str(r.get('counter_name') or r.get('name') or c_id).strip()
            vendor = str(r.get('vendor') or 'Huawei').strip()
            ne = str(r.get('network_element') or r.get('ne') or 'eNodeB').strip()

            if not c_id:
                errors.append(f'Row {idx}: Missing counter_id')
                continue

            NetworkCounterDefinition.objects.update_or_create(
                vendor=vendor,
                network_element=ne,
                counter_id=c_id,
                defaults={
                    'counter_name': c_name,
                    'technology': str(r.get('technology') or r.get('tech') or '4G').strip(),
                    'kpi_code': str(r.get('kpi_code') or r.get('kpi') or '').strip().upper(),
                    'formula_role': str(r.get('formula_role') or 'NUMERATOR').strip().upper(),
                    'description': str(r.get('description') or '').strip(),
                    'is_active': True,
                }
            )
            count += 1

        return JsonResponse({'success': True, 'count': count, 'error_count': len(errors)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def counter_download_template(request):
    """Download sample CSV template for Counter Dictionary import."""
    import csv, io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'vendor', 'network_element', 'counter_id', 'counter_name',
        'technology', 'kpi_code', 'formula_role', 'description'
    ])
    writer.writerow([
        'Huawei', 'eNodeB', 'L.RRC.ConnReq.Att', 'RRC Connection Request Attempts',
        '4G', 'CSSR', 'DENOMINATOR', 'Total RRC connection request attempts'
    ])
    writer.writerow([
        'Huawei', 'eNodeB', 'L.RRC.ConnReq.Succ', 'RRC Connection Request Successes',
        '4G', 'CSSR', 'NUMERATOR', 'Successful RRC connection establishments'
    ])
    writer.writerow([
        'Ericsson', 'gNodeB', 'N.N3GPP.Conn.Att', '5G NR Connection Attempts',
        '5G', 'DATA_ACCESS_SR', 'DENOMINATOR', '5G NR Connection Attempt Count'
    ])

    resp = HttpResponse(output.getvalue(), content_type='text/csv')
    resp['Content-Disposition'] = 'attachment; filename="counters_template.csv"'
    return resp


@login_required
@require_POST
def counter_delete(request, pk):
    get_object_or_404(NetworkCounterDefinition, pk=pk).delete()
    return JsonResponse({'success': True})



